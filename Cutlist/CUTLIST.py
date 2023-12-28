from tkinter.filedialog import askopenfilename 

from tkinter.filedialog import askdirectory

from tkinter import *
from tkinter.ttk import Label
from openpyxl import load_workbook

import pandas as pd
import warnings
warnings.filterwarnings('ignore')
import numpy as np


def openfile():
    # filedialog.askopenfilename()
    entry2.delete(0, END)
    filepath=askopenfilename( filetypes=[("Excel files","*.xlsx")])

    #filepath = askdirectory(parent=troot, initialdir="/",
                            #title='Select a file')
    entry2.insert(0, filepath)
    canvas.create_text(175, 320, text='', font=('Cambria', 17), fill='white')
    troot.update()
    if entry2.get() != '':
        btn2 = Button(canvas, text='Browse', font=(
            'Cambria', 10, 'bold'), command=openfile, bg='#00003f', fg='#FFDF00')
        canvas.create_window(305, 113, anchor='e',
                             window=btn2, height=28, width=55)
    else:
        btn2 = Button(canvas, text='Browse', font=(
            'Cambria', 10, 'bold'), command=openfile, bg='#00003f', fg='#FFDF00')
        canvas.create_window(305, 113, anchor='e',
                             window=btn2, height=28, width=55)
        

troot = Tk()

troot.geometry("350x400")
troot.eval('tk::PlaceWindow . center')
troot.title('CUT LIST')

#print(rgbtohex(r=255, g=255, b=255))
canvas = Canvas(troot, width=500, height=500, bg='#010203')
canvas.pack(fill='both', expand=True)
canvas.create_text(165, 40, text='CUT LIST',
                   font=('Cambria', 16, 'bold'), fill='#FFDF00')
# open folder
canvas.create_text(70, 90, text='Choose the folder',
                   font=('Cambria', 10), fill='#FFDF00')
entry2 = Entry(canvas)
entry2.config(borderwidth=3)
btn2 = Button(canvas, text='Browse', font=('Cambria', 10, 'bold'),
              command=openfile, bg='#00003f', fg='#FFDF00')
canvas.create_window(20, 100, anchor='nw', window=entry2, height=27, width=200)
canvas.create_window(305, 113, anchor='e', window=btn2, height=28, width=55)

def generate():
    if entry2.get() == "":
        folder_warning_label = Label(canvas,foreground='red',background='#010203', text='Please select a Folder.')
        folder_warning_label.place(x=20,y=135)
        folder_warning_label.after(3000, lambda:  folder_warning_label.destroy() )  
    else:
        output = askdirectory( initialdir="/",title='Please select a folder to save the excel file')
        
#Loading excel file and finding sum of qty columns using groupby function and saving into new dataframe ll    
        df = pd.read_excel(entry2.get())
        df['Length']=df['Length'].round(2)
        df['Material'] = df['Material'].str.strip()
        ll =df.groupby(['Type','Material','Length'])['Qty'].sum().to_frame(name='Qty').reset_index()
        pd.set_option('display.max_rows', None)
        
       
#creating separate data frame for king studs   
        kingStud_details = None             
        Ks = ll.loc[ll['Type'] == 'KingStud']
        if len(Ks) != 0:
            Ks['Length'] = Ks['Length'].apply(np.ceil)
            Ks['newMAT'] = Ks['Material'].str.split(' ').str[0] 
            Ks.loc[Ks['Length']<=144, 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '12 SYP #2'
            Ks.loc[Ks['Length']<=144, 'Foot'] =144
            Ks.loc[Ks['Length'].between(144,192), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '16 SYP #2'
            Ks.loc[Ks['Length'].between(144,192), 'Foot'] =16*12
            Ks.loc[Ks['Length'].between(192,240), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '20 SYP #2'
            Ks.loc[Ks['Length'].between(192,240), 'Foot'] =20*12
            Ks.loc[Ks['Length'].between(240,264), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '22 SYP #2'
            Ks.loc[Ks['Length'].between(240,264), 'Foot'] =22*12
            Ks.loc[Ks['Length'].between(264,288), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '24 SYP #2'
            Ks.loc[Ks['Length'].between(264,288), 'Foot'] =24*12
            Ks.loc[Ks['Length'].between(288,312), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '26 SYP #2'
            Ks.loc[Ks['Length'].between(288,312), 'Foot'] =26*12
            Ks.loc[Ks['Length'].between(312,336), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '28 SYP #2'
            Ks.loc[Ks['Length'].between(312,336), 'FFoot'] =28*12
            Ks.loc[Ks['Length'].between(336,360), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '30 SYP #2'
            Ks.loc[Ks['Length'].between(336,360), 'Foot'] =30*12
            Ks.loc[Ks['Length'].between(360,384), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '32 SYP #2'
            Ks.loc[Ks['Length'].between(360,384), 'Foot'] =32*12
            
            #print(Ks)
            
            Ks['newMAT'] = Ks['newMAT'] + 'x' + Ks['Length'].astype(str)
            Ks2 = Ks.loc[~Ks['Material'].str.contains('PT')]
            kingStud_details = (Ks2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            kingStud_details = kingStud_details.rename(columns = {'newMAT':'Material'})
            #print(kingStud_details)
            kingStud_details['Qty'] = (kingStud_details['Qty'] * kingStud_details['Length'])/kingStud_details['Foot']
            kingStud_details['Qty'] = kingStud_details['Qty'].apply(np.ceil)
            #print(kingStud_details," king stud")
        
        
#creating separate data frame for PT king studs
        kingStud_details_pt = None           
        Ks = ll.loc[ll['Type'] == 'KingStud']
        if len(Ks) != 0:
            Ks['Length'] = Ks['Length'].apply(np.ceil)
            Ks['newMAT'] = Ks['Material'].str.split(' ').str[0]
            Ks.loc[Ks['Length']<=144, 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '12 PT SYP #2'
            Ks.loc[Ks['Length']<=144, 'Foot'] =144
            Ks.loc[Ks['Length'].between(144,192), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '16 PT SYP #2'
            Ks.loc[Ks['Length'].between(144,192), 'Foot'] =16*12
            Ks.loc[Ks['Length'].between(192,240), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '20 PT SYP #2'
            Ks.loc[Ks['Length'].between(192,240), 'Foot'] =20*12
            Ks.loc[Ks['Length'].between(240,264), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '22 PT SYP #2'
            Ks.loc[Ks['Length'].between(240,264), 'Foot'] =22*12
            Ks.loc[Ks['Length'].between(264,288), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '24 PT SYP #2'
            Ks.loc[Ks['Length'].between(264,288), 'Foot'] =24*12
            Ks.loc[Ks['Length'].between(288,312), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '26 PT SYP #2'
            Ks.loc[Ks['Length'].between(288,312), 'Foot'] =26*12
            Ks.loc[Ks['Length'].between(312,336), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '28 PT SYP #2'
            Ks.loc[Ks['Length'].between(312,336), 'FFoot'] =28*12
            Ks.loc[Ks['Length'].between(336,360), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '30 PT SYP #2'
            Ks.loc[Ks['Length'].between(336,360), 'Foot'] =30*12
            Ks.loc[Ks['Length'].between(360,384), 'From Fishbowl Inventory'] = Ks['newMAT'] + 'x' + '32 PT SYP #2'
            Ks.loc[Ks['Length'].between(360,384), 'Foot'] =32*12
            Ks['newMAT'] = Ks['newMAT'] + 'x' + Ks['Length'].astype(str)
            Ks1 = Ks.loc[Ks['Material'].str.contains('PT')]
            kingStud_details_pt = (Ks1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            kingStud_details_pt = kingStud_details_pt.rename(columns = {'newMAT':'Material'})
            kingStud_details_pt['Qty'] = (kingStud_details_pt['Qty'] * kingStud_details_pt['Length'])/kingStud_details_pt['Foot']
            kingStud_details_pt['Qty'] = kingStud_details_pt['Qty'].apply(np.ceil)
            #print(kingStud_details_pt)
        
#creating separate data frame for studs   
        Stud_details = None                
        S = ll.loc[ll['Type'] == 'Stud']
        if len(S) != 0:
            S['Length'] = S['Length'].apply(np.ceil)
            S['newMAT'] = S['Material'].str.split(' ').str[0]
            S.loc[S['Length']<=144, 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '12 SYP #2'
            S.loc[S['Length']<=144, 'Foot'] =144
            S.loc[S['Length'].between(144,192), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '16 SYP #2'
            S.loc[S['Length'].between(144,192), 'Foot'] =16*12
            S.loc[S['Length'].between(192,240), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '20 SYP #2'
            S.loc[S['Length'].between(192,240), 'Foot'] =20*12
            S.loc[S['Length'].between(240,264), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '22 SYP #2'
            S.loc[S['Length'].between(240,264), 'Foot'] =22*12
            S.loc[S['Length'].between(264,288), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '24 SYP #2'
            S.loc[S['Length'].between(264,288), 'Foot'] =24*12
            S.loc[S['Length'].between(288,312), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '26 SYP #2'
            S.loc[S['Length'].between(288,312), 'Foot'] =26*12
            S.loc[S['Length'].between(312,336), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '28 SYP #2'
            S.loc[S['Length'].between(312,336), 'Foot'] =28*12
            S.loc[S['Length'].between(336,360), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '30 SYP #2'
            S.loc[S['Length'].between(336,360), 'Foot'] =30*12
            S.loc[S['Length'].between(360,384), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '32 SYP #2'
            S.loc[S['Length'].between(360,384), 'Foot'] =32*12
            S['newMAT'] = S['newMAT'] + 'x' + S['Length'].astype(str)
            S2 = S.loc[~S['Material'].str.contains('PT')]
            Stud_details = (S2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            Stud_details = Stud_details.rename(columns={'newMAT':'Material'})
            Stud_details['Qty'] = (Stud_details['Qty'] * Stud_details['Length'])/Stud_details['Foot']
            Stud_details['Qty'] = Stud_details['Qty'].apply(np.ceil)
            #print(Stud_details)
            #print(Stud_details, "stud ")

#creating separate data frame for PT studs                   
        Stud_details_pt = None
        S = ll.loc[ll['Type'] == 'Stud']
        if len(S) != 0:
            S['Length'] = S['Length'].apply(np.ceil)
            S['newMAT'] = S['Material'].str.split(' ').str[0]
            S.loc[S['Length']<=144, 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '12 PT SYP #2'
            S.loc[S['Length']<=144, 'Foot'] =144
            S.loc[S['Length'].between(144,192), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '16 PT SYP #2'
            S.loc[S['Length'].between(144,192), 'Foot'] =16*12
            S.loc[S['Length'].between(192,240), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '20 PT SYP #2'
            S.loc[S['Length'].between(192,240), 'Foot'] =20*12
            S.loc[S['Length'].between(240,264), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '22 PT SYP #2'
            S.loc[S['Length'].between(240,264), 'Foot'] =22*12
            S.loc[S['Length'].between(264,288), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '24 PT SYP #2'
            S.loc[S['Length'].between(264,288), 'Foot'] =24*12
            S.loc[S['Length'].between(288,312), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '26 PT SYP #2'
            S.loc[S['Length'].between(288,312), 'Foot'] =26*12
            S.loc[S['Length'].between(312,336), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '28 PT SYP #2'
            S.loc[S['Length'].between(312,336), 'Foot'] =28*12
            S.loc[S['Length'].between(336,360), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '30 PT SYP #2'
            S.loc[S['Length'].between(336,360), 'Foot'] =30*12
            S.loc[S['Length'].between(360,384), 'From Fishbowl Inventory'] = S['newMAT'] + 'x' + '32 PT SYP #2'  
            S.loc[S['Length'].between(360,384), 'Foot'] =32*12
            S['newMAT'] = S['newMAT'] + 'x' + S['Length'].astype(str)
            S1 = S.loc[S['Material'].str.contains('PT')]
            Stud_details_pt = (S1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            Stud_details_pt = Stud_details_pt.rename(columns={'newMAT':'Material'})
            Stud_details_pt['Qty'] = (Stud_details_pt['Qty'] * Stud_details_pt['Length'])/Stud_details_pt['Foot']
            Stud_details_pt['Qty'] = Stud_details_pt['Qty'].apply(np.ceil)
            #print(Stud_details_pt)
        
#creating separate data frame for post          
        Post_details = None                 
        P = ll.loc[ll['Type'] == 'Post']
        if len(P) != 0:
            P['newMAT'] = P['Material'].str.split(' ').str[0]
            P.loc[P['Length']<=144, 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '12 SYP #2'
            P.loc[P['Length']<=144, 'Foot'] =12*12
            P.loc[P['Length'].between(144,192), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '16 SYP #2'
            P.loc[P['Length'].between(144,192), 'Foot'] =16*12
            P.loc[P['Length'].between(192,240), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '20 SYP #2'
            P.loc[P['Length'].between(192,240), 'Foot'] =20*12
            P.loc[P['Length'].between(240,264), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '22 SYP #2'
            P.loc[P['Length'].between(240,264), 'Foot'] =22*12
            P.loc[P['Length'].between(264,288), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '24 SYP #2'
            P.loc[P['Length'].between(264,288), 'Foot'] =24*12
            P.loc[P['Length'].between(288,312), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '26 SYP #2'
            P.loc[P['Length'].between(288,312), 'Foot'] =26*12
            P.loc[P['Length'].between(312,336), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '28 SYP #2'
            P.loc[P['Length'].between(312,336), 'Foot'] =28*12
            P.loc[P['Length'].between(336,360), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '30 SYP #2'
            P.loc[P['Length'].between(336,360), 'Foot'] =30*12
            P.loc[P['Length'].between(360,384), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '32 SYP #2'
            P.loc[P['Length'].between(360,384), 'Foot'] =32*12
            P['newMAT'] = P['newMAT'] + 'x' + P['Length'].astype(str)
            P1 = P.loc[(~P['Material'].str.contains('LVL')) & (~P['Material'].str.contains('PT'))]
            Post_details = (P1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            Post_details  = Post_details.rename(columns={'newMAT':'Material'})
            Post_details['Qty'] = (Post_details['Qty'] * Post_details['Length'])/Post_details['Foot']
            Post_details['Qty'] = Post_details['Qty'].apply(np.ceil)
            #print(Post_details)
        
#creating separate data frame for PT post    
        Post_details_pt = None                       
        P = ll.loc[ll['Type'] == 'Post']
        if len(P) != 0:
            P['newMAT'] = P['Material'].str.split(' ').str[0]
            P.loc[P['Length']<=144, 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '12 PT SYP #2'
            P.loc[P['Length']<=144, 'Foot'] =12*12
            P.loc[P['Length'].between(144,192), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '16 PT SYP #2'
            P.loc[P['Length'].between(144,192), 'Foot'] =16*12
            P.loc[P['Length'].between(192,240), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '20 PT SYP #2'
            P.loc[P['Length'].between(192,240), 'Foot'] =20*12
            P.loc[P['Length'].between(240,264), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '22 PT SYP #2'
            P.loc[P['Length'].between(240,264), 'Foot'] =22*12
            P.loc[P['Length'].between(264,288), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '24 PT SYP #2'
            P.loc[P['Length'].between(264,288), 'Foot'] =24*12
            P.loc[P['Length'].between(288,312), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '26 PT SYP #2'
            P.loc[P['Length'].between(288,312), 'Foot'] =26*12
            P.loc[P['Length'].between(312,336), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '28 PT SYP #2'
            P.loc[P['Length'].between(312,336), 'Foot'] =28*12
            P.loc[P['Length'].between(336,360), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '30 PT SYP #2'
            P.loc[P['Length'].between(336,360), 'Foot'] =30*12
            P.loc[P['Length'].between(360,384), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '32 PT SYP #2'
            P.loc[P['Length'].between(360,384), 'Foot'] =32*12
            P['newMAT'] = P['newMAT'] + 'x' + P['Length'].astype(str)
            P2 = P.loc[P['Material'].str.contains('PT')]
            Post_details_pt = (P2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            Post_details_pt = Post_details_pt.rename(columns={'newMAT':'Material'})
            Post_details_pt['Qty'] = (Post_details_pt['Qty'] * Post_details_pt['Length'])/Post_details_pt['Foot']
            Post_details_pt['Qty'] = Post_details_pt['Qty'].apply(np.ceil)
           #print(Post_details_pt)
        
#creating separate data frame for PT post  
        Post_details_lvl = None                                 
        P = ll.loc[ll['Type'] == 'Post']
        if len(P) != 0:
            P['newMAT'] = P['Material'].str.split(' ').str[0]
            P.loc[P['Length']<=144, 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '12 LVL SYP #2'
            P.loc[P['Length']<=144, 'Foot'] =12*12
            P.loc[P['Length'].between(144,192), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '16 LVL SYP #2'
            P.loc[P['Length'].between(144,192), 'Foot'] =16*12
            P.loc[P['Length'].between(192,240), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '20 LVL SYP #2'
            P.loc[P['Length'].between(192,240), 'Foot'] =20*12
            P.loc[P['Length'].between(240,264), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '22 LVL SYP #2'
            P.loc[P['Length'].between(240,264), 'Foot'] =22*12
            P.loc[P['Length'].between(264,288), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '24 LVL SYP #2'
            P.loc[P['Length'].between(264,288), 'Foot'] =24*12
            P.loc[P['Length'].between(288,312), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '26 LVL SYP #2'
            P.loc[P['Length'].between(288,312), 'Foot'] =26*12
            P.loc[P['Length'].between(312,336), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '28 LVL SYP #2'
            P.loc[P['Length'].between(312,336), 'Foot'] =28*12
            P.loc[P['Length'].between(336,360), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '30 LVL SYP #2'
            P.loc[P['Length'].between(336,360), 'Foot'] =30*12
            P.loc[P['Length'].between(360,384), 'From Fishbowl Inventory'] = P['newMAT'] + 'x' + '32 LVL SYP #2'
            P.loc[P['Length'].between(360,384), 'Foot'] =32*12
            P['newMAT'] = P['newMAT'] + 'x' + P['Length'].astype(str)
            P2 = P.loc[P['Material'].str.contains('LVL')]
            Post_details_lvl = (P2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            Post_details_lvl = Post_details_lvl.rename(columns={'newMAT':'Material'})
            Post_details_lvl['Qty'] = (Post_details_lvl['Qty'] * Post_details_lvl['Length'])/Post_details_lvl['Foot']
            Post_details_lvl['Qty'] = Post_details_lvl['Qty'].apply(np.ceil)
            #print(Post_details_lvl)
        
        
        JackOverOpening_details = None
#creating separate data frame for SFJackOverOpening                                           
        JOO = ll.loc[ll['Type'] == 'SFJackOverOpening']
        print(len(JOO),"the length of the jooo")
        if len(JOO) != 0:
            JOO['Length'] = JOO['Length'].apply(np.ceil)
            JOO['newMAT'] = JOO['Material'].str.split(' ').str[0]
            JOO.loc[JOO['Length']<=144, 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '12 SYP #2'
            JOO.loc[JOO['Length']<=144, 'Foot'] = 144
            JOO.loc[JOO['Length'].between(144,192), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '16 SYP #2'
            JOO.loc[JOO['Length'].between(144,192), 'Foot'] =16*12
            JOO.loc[JOO['Length'].between(192,240), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '20 SYP #2'
            JOO.loc[JOO['Length'].between(192,240), 'Foot'] =20*12
            JOO.loc[JOO['Length'].between(240,264), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '22 SYP #2'
            JOO.loc[JOO['Length'].between(240,264), 'Foot'] =22*12
            JOO.loc[JOO['Length'].between(264,288), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '24 SYP #2'
            JOO.loc[JOO['Length'].between(264,288), 'Foot'] =24*12
            JOO.loc[JOO['Length'].between(288,312), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '26 SYP #2'
            JOO.loc[JOO['Length'].between(288,312), 'Foot'] =26*12
            JOO.loc[JOO['Length'].between(312,336), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '28 SYP #2'
            JOO.loc[JOO['Length'].between(312,336), 'Foot'] =28*12
            JOO.loc[JOO['Length'].between(336,360), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '30 SYP #2'
            JOO.loc[JOO['Length'].between(336,360), 'Foot'] =30*12
            JOO.loc[JOO['Length'].between(360,384), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '32 SYP #2'
            JOO.loc[JOO['Length'].between(360,384), 'Foot'] =32*12
            
            JOO['newMAT'] = JOO['newMAT'] + 'x' + JOO['Length'].astype(str)
            JOO2 = JOO.loc[~JOO['Material'].str.contains('PT')]
            JackOverOpening_details = (JOO2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            JackOverOpening_details = JackOverOpening_details.rename(columns={'newMAT':'Material'})
            JackOverOpening_details['Qty'] = (JackOverOpening_details['Qty'] * JackOverOpening_details['Length'])/JackOverOpening_details['Foot']
            JackOverOpening_details['Qty'] = JackOverOpening_details['Qty'].apply(np.ceil)
            print(JackOverOpening_details,"JAckover")
        #print(JackOverOpening_details)
        
#creating separate data frame for PT SFJackOverOpening          
        JackOverOpening_details_pt = None
        JOO = ll.loc[ll['Type'] == 'SFJackOverOpening']
        if len(JOO) != 0:
            JOO['Length'] = JOO['Length'].apply(np.ceil)
            JOO['newMAT'] = JOO['Material'].str.split(' ').str[0]
            JOO.loc[JOO['Length']<=144, 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '12 PT SYP #2'
            JOO.loc[JOO['Length']<=144, 'Foot'] = 14
            JOO.loc[JOO['Length'].between(144,192), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '16 PT SYP #2'
            JOO.loc[JOO['Length'].between(144,192), 'Foot'] =16*12
            JOO.loc[JOO['Length'].between(192,240), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '20 PT SYP #2'
            JOO.loc[JOO['Length'].between(192,240), 'Foot'] =20*12
            JOO.loc[JOO['Length'].between(240,264), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '22 PT SYP #2'
            JOO.loc[JOO['Length'].between(240,264), 'Foot'] =22*12
            JOO.loc[JOO['Length'].between(264,288), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '24 PT SYP #2'
            JOO.loc[JOO['Length'].between(264,288), 'Foot'] =24*12
            JOO.loc[JOO['Length'].between(288,312), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '26 PT SYP #2'
            JOO.loc[JOO['Length'].between(288,312), 'Foot'] =26*12
            JOO.loc[JOO['Length'].between(312,336), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '28 PT SYP #2'
            JOO.loc[JOO['Length'].between(312,336), 'Foot'] =28*12
            JOO.loc[JOO['Length'].between(336,360), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '30 PT SYP #2'
            JOO.loc[JOO['Length'].between(336,360), 'Foot'] =30*12
            JOO.loc[JOO['Length'].between(360,384), 'From Fishbowl Inventory'] = JOO['newMAT'] + 'x' + '32 PT SYP #2'
            JOO.loc[JOO['Length'].between(360,384), 'Foot'] =32*12
            JOO['newMAT'] = JOO['newMAT'] + 'x' + JOO['Length'].astype(str)
            JOO1 = JOO.loc[JOO['Material'].str.contains('PT')]
            JackOverOpening_details_pt = (JOO1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            JackOverOpening_details_pt = JackOverOpening_details_pt.rename(columns={'newMAT':'Material'})
            JackOverOpening_details_pt['Qty'] = (JackOverOpening_details_pt['Qty'] * JackOverOpening_details_pt['Length'])/JackOverOpening_details_pt['Foot']
            JackOverOpening_details_pt['Qty'] = JackOverOpening_details_pt['Qty'].apply(np.ceil)
            #print(JackOverOpening_details_pt)
        
        
        
#creating separate data frame for SFJackUnderOpening     
        JackUnderOpening_details = None                                        
        JUO = ll.loc[ll['Type'] == 'SFJackUnderOpening']
        if len(JUO) != 0:
            #print(JUO,"EXECUTED ")
            JUO['Length'] = JUO['Length'].apply(np.ceil)
            JUO['newMAT'] = JUO['Material'].str.split(' ').str[0]
            JUO.loc[JUO['Length']<=144, 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '12 SYP #2'
            JUO.loc[JUO['Length']<=144, 'Foot'] =12*12
            JUO.loc[JUO['Length'].between(144,192), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '16 SYP #2'
            JUO.loc[JUO['Length'].between(144,192), 'Foot'] =16*12
            JUO.loc[JUO['Length'].between(192,240), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '20 SYP #2'
            JUO.loc[JUO['Length'].between(192,240), 'Foot'] =20*12
            JUO.loc[JUO['Length'].between(240,264), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '22 SYP #2'
            JUO.loc[JUO['Length'].between(240,264), 'Foot'] =22*12
            JUO.loc[JUO['Length'].between(264,288), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '24 SYP #2'
            JUO.loc[JUO['Length'].between(264,288), 'Foot'] =24*12
            JUO.loc[JUO['Length'].between(288,312), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '26 SYP #2'
            JUO.loc[JUO['Length'].between(288,312), 'Foot'] =26*12
            JUO.loc[JUO['Length'].between(312,336), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '28 SYP #2'
            JUO.loc[JUO['Length'].between(312,336), 'Foot'] =28*12
            JUO.loc[JUO['Length'].between(336,360), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '30 SYP #2'
            JUO.loc[JUO['Length'].between(336,360), 'Foot'] =30*12
            JUO.loc[JUO['Length'].between(360,384), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '32 SYP #2'
            JUO.loc[JUO['Length'].between(360,384), 'Foot'] =32*12
            JUO['newMAT'] = JUO['newMAT'] + 'x' + JUO['Length'].astype(str)
            JUO1 = JUO.loc[~JUO['Material'].str.contains('PT')]
            JackUnderOpening_details = (JUO1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            JackUnderOpening_details = JackUnderOpening_details.rename(columns={'newMAT':'Material'})
            JackUnderOpening_details['Qty'] = (JackUnderOpening_details['Qty'] * JackUnderOpening_details['Length'])/JackUnderOpening_details['Foot']
            JackUnderOpening_details['Qty'] = JackUnderOpening_details['Qty'].apply(np.ceil)
    
            print(JackUnderOpening_details," JUO")
        #print(JackUnderOpening_details)
        
#creating separate data frame for PT SFJackUnderOpening        
        JackUnderOpening_details_pt = None                                     
        JUO = ll.loc[ll['Type'] == 'SFJackUnderOpening']
        if len(JUO) != 0:
            JUO['Length'] = JUO['Length'].apply(np.ceil)
            JUO['newMAT'] = JUO['Material'].str.split(' ').str[0]
            JUO.loc[JUO['Length']<=144, 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '12 PT SYP #2'
            JUO.loc[JUO['Length']<=144, 'Foot'] =12*12
            JUO.loc[JUO['Length'].between(144,192), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '16 PT SYP #2'
            JUO.loc[JUO['Length'].between(144,192), 'Foot'] =16*12
            JUO.loc[JUO['Length'].between(192,240), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '20 PT SYP #2'
            JUO.loc[JUO['Length'].between(192,240), 'Foot'] =20*12
            JUO.loc[JUO['Length'].between(240,264), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '22 PT SYP #2'
            JUO.loc[JUO['Length'].between(240,264), 'Foot'] =22*12
            JUO.loc[JUO['Length'].between(264,288), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '24 PT SYP #2'
            JUO.loc[JUO['Length'].between(264,288), 'Foot'] =24*12
            JUO.loc[JUO['Length'].between(288,312), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '26 PT SYP #2'
            JUO.loc[JUO['Length'].between(288,312), 'Foot'] =26*12
            JUO.loc[JUO['Length'].between(312,336), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '28 PT SYP #2'
            JUO.loc[JUO['Length'].between(312,336), 'Foot'] =28*12
            JUO.loc[JUO['Length'].between(336,360), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '30 PT SYP #2'
            JUO.loc[JUO['Length'].between(336,360), 'Foot'] =30*12
            JUO.loc[JUO['Length'].between(360,384), 'From Fishbowl Inventory'] = JUO['newMAT'] + 'x' + '32 PT SYP #2'
            JUO.loc[JUO['Length'].between(360,384), 'Foot'] =32*12
            JUO['newMAT'] = JUO['newMAT'] + 'x' + JUO['Length'].astype(str)
            JUO2 = JUO.loc[JUO['Material'].str.contains('PT')]
            JackUnderOpening_details_pt = (JUO2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            JackUnderOpening_details_pt = JackUnderOpening_details_pt.rename(columns={'newMAT':'Material'})
            JackUnderOpening_details_pt['Qty'] = (JackUnderOpening_details_pt['Qty'] * JackUnderOpening_details_pt['Length'])/JackUnderOpening_details_pt['Foot']
            JackUnderOpening_details_pt['Qty'] = JackUnderOpening_details_pt['Qty'].apply(np.ceil)
            JackUnderOpening_details_pt
        
        
#creating separate data frame for SFJackUnderOpening
        cripple_details = None                                             
        Cr = ll.loc[ll['Type'] == 'Cripple']
        if len(Cr)!= 0 :
            Cr['Length'] = Cr['Length'].apply(np.ceil)
            Cr['newMAT'] = Cr['Material'].str.split(' ').str[0]
            Cr.loc[Cr['Length']<=144, 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '12 SYP #2'
            Cr.loc[Cr['Length']<=144, 'Foot'] =12*12
            Cr.loc[Cr['Length'].between(144,192), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '16 SYP #2'
            Cr.loc[Cr['Length'].between(144,192), 'Foot'] =16*12
            Cr.loc[Cr['Length'].between(192,240), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '20 SYP #2'
            Cr.loc[Cr['Length'].between(192,240), 'Foot'] =20*12
            Cr.loc[Cr['Length'].between(240,264), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '22 SYP #2'
            Cr.loc[Cr['Length'].between(240,264), 'Foot'] =22*12
            Cr.loc[Cr['Length'].between(264,288), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '24 SYP #2'
            Cr.loc[Cr['Length'].between(264,288), 'Foot'] =24*12
            Cr.loc[Cr['Length'].between(288,312), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '26 SYP #2'
            Cr.loc[Cr['Length'].between(288,312), 'Foot'] =26*12
            Cr.loc[Cr['Length'].between(312,336), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '28 SYP #2'
            Cr.loc[Cr['Length'].between(312,336), 'Foot'] =28*12
            Cr.loc[Cr['Length'].between(336,360), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '30 SYP #2'
            Cr.loc[Cr['Length'].between(336,360), 'Foot'] =30*12
            Cr.loc[Cr['Length'].between(360,384), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '32 SYP #2'
            Cr.loc[Cr['Length'].between(360,384), 'Foot'] =32*12
            Cr['newMAT'] = Cr['newMAT'] + 'x' + Cr['Length'].astype(str)
            Cr1 = Cr.loc[~Cr['Material'].str.contains('PT')]
            cripple_details = (Cr1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            cripple_details = cripple_details.rename(columns={'newMAT':'Material'})
            cripple_details['Qty'] = (cripple_details['Qty'] * cripple_details['Length'])/cripple_details['Foot']
            cripple_details['Qty'] = cripple_details['Qty'].apply(np.ceil)
    
            print(cripple_details,"cripple")
        #print(JackUnderOpening_details)
        
#creating separate data frame for PT SFJackUnderOpening
        cripple_details_pt = None                                             
        Cr = ll.loc[ll['Type'] == 'Cripple']
        if len(Cr) != 0:
            Cr['Length'] = Cr['Length'].apply(np.ceil)
            Cr['newMAT'] = Cr['Material'].str.split(' ').str[0]
            Cr.loc[Cr['Length']<=144, 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '12 PT SYP #2'
            Cr.loc[Cr['Length']<=144, 'Foot'] =12*12
            Cr.loc[Cr['Length'].between(144,192), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '16 PT SYP #2'
            Cr.loc[Cr['Length'].between(144,192), 'Foot'] =16*12
            Cr.loc[Cr['Length'].between(192,240), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '20 PT SYP #2'
            Cr.loc[Cr['Length'].between(192,240), 'Foot'] =20*12
            Cr.loc[Cr['Length'].between(240,264), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '22 PT SYP #2'
            Cr.loc[Cr['Length'].between(240,264), 'Foot'] =22*12
            Cr.loc[Cr['Length'].between(264,288), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '24 PT SYP #2'
            Cr.loc[Cr['Length'].between(264,288), 'Foot'] =24*12
            Cr.loc[Cr['Length'].between(288,312), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '26 PT SYP #2'
            Cr.loc[Cr['Length'].between(288,312), 'Foot'] =26*12
            Cr.loc[Cr['Length'].between(312,336), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '28 PT SYP #2'
            Cr.loc[Cr['Length'].between(312,336), 'Foot'] =28*12
            Cr.loc[Cr['Length'].between(336,360), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '30 PT SYP #2'
            Cr.loc[Cr['Length'].between(336,360), 'Foot'] =30*12
            Cr.loc[Cr['Length'].between(360,384), 'From Fishbowl Inventory'] = Cr['newMAT'] + 'x' + '32 PT SYP #2'
            Cr.loc[Cr['Length'].between(360,384), 'Foot'] =32*12
            Cr['newMAT'] = Cr['newMAT'] + 'x' + Cr['Length'].astype(str)
            Cr2 = Cr.loc[Cr['Material'].str.contains('PT')]
            cripple_details_pt = (Cr2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            cripple_details_pt = cripple_details_pt.rename(columns={'newMAT':'Material'})
            cripple_details_pt['Qty'] = (cripple_details_pt['Qty'] * cripple_details_pt['Length'])/cripple_details_pt['Foot']
            cripple_details_pt['Qty'] = cripple_details_pt['Qty'].apply(np.ceil)
            cripple_details_pt
        
#creating separate data frame for SFSupportingBeam                                                     
        Beam_details = None
        SB = ll.loc[ll['Type'] == 'SFSupportingBeam']
        if len(SB) != 0:
            SB['Length'] = SB['Length'].apply(np.ceil)
            SB['newMAT'] = SB['Material'].str.split(' ').str[0]
            SB.loc[SB['Length']<=144, 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '12 SYP #2'
            SB.loc[SB['Length']<=144, 'Foot'] =12*12
            SB.loc[SB['Length'].between(144,192), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '16 SYP #2'
            SB.loc[SB['Length'].between(144,192), 'Foot'] =16*12
            SB.loc[SB['Length'].between(192,240), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '20 SYP #2'
            SB.loc[SB['Length'].between(192,240), 'Foot'] =20*12
            SB.loc[SB['Length'].between(240,264), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '22 SYP #2'
            SB.loc[SB['Length'].between(240,264), 'Foot'] =22*12
            SB.loc[SB['Length'].between(264,288), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '24 SYP #2'
            SB.loc[SB['Length'].between(264,288), 'Foot'] =24*12
            SB.loc[SB['Length'].between(288,312), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '26 SYP #2'
            SB.loc[SB['Length'].between(288,312), 'Foot'] =26*12
            SB.loc[SB['Length'].between(312,336), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '28 SYP #2'
            SB.loc[SB['Length'].between(312,336), 'Foot'] =28*12
            SB.loc[SB['Length'].between(336,360), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '30 SYP #2'
            SB.loc[SB['Length'].between(336,360), 'Foot'] =30*12
            SB.loc[SB['Length'].between(360,384), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '32 SYP #2'
            SB.loc[SB['Length'].between(360,384), 'Foot'] =32*12
            SB['newMAT'] = SB['newMAT'] + 'x' + SB['Length'].astype(str)
            SB1 = SB.loc[~SB['Material'].str.contains('PT')]
            Beam_details = (SB1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            Beam_details = Beam_details.rename(columns={'newMAT':'Material'})
            Beam_details['Qty'] = (Beam_details['Qty'] * Beam_details['Length'])/Beam_details['Foot']
            Beam_details['Qty'] = Beam_details['Qty'].apply(np.ceil)
            #print(Beam_details)

#creating separate data frame for PT SFSupportingBeam
        Beam_details_pt = None                                                    
        SB = ll.loc[ll['Type'] == 'SFSupportingBeam']
        if len(SB) != 0:    
            SB['Length'] = SB['Length'].apply(np.ceil)
            SB['newMAT'] = SB['Material'].str.split(' ').str[0]
            SB.loc[SB['Length']<=144, 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '12 PT SYP #2'
            SB.loc[SB['Length']<=144, 'Foot'] =12*12
            SB.loc[SB['Length'].between(144,192), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '16 PT SYP #2'
            SB.loc[SB['Length'].between(144,192), 'Foot'] =16*12
            SB.loc[SB['Length'].between(192,240), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '20 PT SYP #2'
            SB.loc[SB['Length'].between(192,240), 'Foot'] =20*12
            SB.loc[SB['Length'].between(240,264), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '22 PT SYP #2'
            SB.loc[SB['Length'].between(240,264), 'Foot'] =22*12
            SB.loc[SB['Length'].between(264,288), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '24 PT SYP #2'
            SB.loc[SB['Length'].between(264,288), 'Foot'] =24*12
            SB.loc[SB['Length'].between(288,312), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '26 PT SYP #2'
            SB.loc[SB['Length'].between(288,312), 'Foot'] =26*12
            SB.loc[SB['Length'].between(312,336), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '28 PT SYP #2'
            SB.loc[SB['Length'].between(312,336), 'Foot'] =28*12
            SB.loc[SB['Length'].between(336,360), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '30 PT SYP #2'
            SB.loc[SB['Length'].between(336,360), 'Foot'] =30*12
            SB.loc[SB['Length'].between(360,384), 'From Fishbowl Inventory'] = SB['newMAT'] + 'x' + '32 PT SYP #2'
            SB.loc[SB['Length'].between(360,384), 'Foot'] =32*12
            SB['newMAT'] = SB['newMAT'] + 'x' + SB['Length'].astype(str)
            SB2 = SB.loc[SB['Material'].str.contains('PT')]
            Beam_details_pt = (SB2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            Beam_details_pt = Beam_details_pt.rename(columns={'newMAT':'Material'})
            Beam_details_pt['Qty'] = (Beam_details_pt['Qty'] * Beam_details_pt['Length'])/Beam_details_pt['Foot']
            Beam_details_pt['Qty'] = Beam_details_pt['Qty'].apply(np.ceil)
            Beam_details_pt
        
#creating separate data frame for SFStudLeft
        LS_details = None                                                                     
        LS = ll.loc[ll['Type'] == 'SFStudLeft']
        if len(LS) != 0:
            LS['Length'] = LS['Length'].apply(np.ceil)
            LS['newMAT'] = LS['Material'].str.split(' ').str[0]
            LS.loc[LS['Length']<=144, 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '12 SYP #2'
            LS.loc[LS['Length']<=144, 'Foot'] =12*12
            LS.loc[LS['Length'].between(144,192), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '16 SYP #2'
            LS.loc[LS['Length'].between(144,192), 'Foot'] =16*12
            LS.loc[LS['Length'].between(192,240), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '20 SYP #2'
            LS.loc[LS['Length'].between(192,240), 'Foot'] =20*12
            LS.loc[LS['Length'].between(240,264), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '22 SYP #2'
            LS.loc[LS['Length'].between(240,264), 'Foot'] =22*12
            LS.loc[LS['Length'].between(264,288), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '24 SYP #2'
            LS.loc[LS['Length'].between(264,288), 'Foot'] =24*12
            LS.loc[LS['Length'].between(288,312), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '26 SYP #2'
            LS.loc[LS['Length'].between(288,312), 'Foot'] =26*12
            LS.loc[LS['Length'].between(312,336), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '28 SYP #2'
            LS.loc[LS['Length'].between(312,336), 'Foot'] =28*12
            LS.loc[LS['Length'].between(336,360), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '30 SYP #2'
            LS.loc[LS['Length'].between(336,360), 'Foot'] =30*12
            LS.loc[LS['Length'].between(360,384), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '32 SYP #2'
            LS.loc[LS['Length'].between(360,384), 'Foot'] =32*12
            LS['newMAT'] = LS['newMAT'] + 'x' + LS['Length'].astype(str)
            LS1 = LS.loc[~LS['Material'].str.contains('PT')]
            LS_details = (LS1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            LS_details = LS_details.rename(columns={'newMAT':'Material'})
            LS_details['Qty'] = (LS_details['Qty'] * LS_details['Length'])/LS_details['Foot']
            LS_details['Qty'] = LS_details['Qty'].apply(np.ceil)
            #print(LS_details, "left stud")
        
#creating separate data frame for PT SFStudLeft  
        LS_details_pt = None                                                           
        LS = ll.loc[ll['Type'] == 'SFStudLeft']
        if len(LS) != 0:
            LS['Length'] = LS['Length'].apply(np.ceil)
            LS['newMAT'] = LS['Material'].str.split(' ').str[0]
            LS.loc[LS['Length']<=144, 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '12 PT SYP #2'
            LS.loc[LS['Length']<=144, 'Foot'] =12*12
            LS.loc[LS['Length'].between(144,192), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '16 PT SYP #2'
            LS.loc[LS['Length'].between(144,192), 'Foot'] =16*12
            LS.loc[LS['Length'].between(192,240), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '20 PT SYP #2'
            LS.loc[LS['Length'].between(192,240), 'Foot'] =20*12
            LS.loc[LS['Length'].between(240,264), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '22 PT SYP #2'
            LS.loc[LS['Length'].between(240,264), 'Foot'] =22*12
            LS.loc[LS['Length'].between(264,288), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '24 PT SYP #2'
            LS.loc[LS['Length'].between(264,288), 'Foot'] =24*12
            LS.loc[LS['Length'].between(288,312), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '26 PT SYP #2'
            LS.loc[LS['Length'].between(288,312), 'Foot'] =26*12
            LS.loc[LS['Length'].between(312,336), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '28 PT SYP #2'
            LS.loc[LS['Length'].between(312,336), 'Foot'] =28*12
            LS.loc[LS['Length'].between(336,360), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '30 PT SYP #2'
            LS.loc[LS['Length'].between(336,360), 'Foot'] =30*12
            LS.loc[LS['Length'].between(360,384), 'From Fishbowl Inventory'] = LS['newMAT'] + 'x' + '32 PT SYP #2'
            LS.loc[LS['Length'].between(360,384), 'Foot'] =32*12
            LS['newMAT'] = LS['newMAT'] + 'x' + LS['Length'].astype(str)
            LS2 = LS.loc[LS['Material'].str.contains('PT')]
            LS_details_pt = (LS2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            LS_details_pt = LS_details_pt.rename(columns={'newMAT':'Material'})
            LS_details_pt['Qty'] = (LS_details_pt['Qty'] * LS_details_pt['Length'])/LS_details_pt['Foot']
            LS_details_pt['Qty'] = LS_details_pt['Qty'].apply(np.ceil)
            LS_details_pt
        
#creating separate data frame for SFStudRight                                                                     
        RS_details = None
        RS = ll.loc[ll['Type'] == 'SFStudRight']
        if len(RS) != 0:
            RS['Length'] = RS['Length'].apply(np.ceil)
            RS['newMAT'] = RS['Material'].str.split(' ').str[0]
            RS.loc[RS['Length']<=144, 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '12 SYP #2'
            RS.loc[RS['Length']<=144, 'Foot'] =12*12
            RS.loc[RS['Length'].between(144,192), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '16 SYP #2'
            RS.loc[RS['Length'].between(144,192), 'Foot'] =16*12
            RS.loc[RS['Length'].between(192,240), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '20 SYP #2'
            RS.loc[RS['Length'].between(192,240), 'Foot'] =20*12
            RS.loc[RS['Length'].between(240,264), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '22 SYP #2'
            RS.loc[RS['Length'].between(240,264), 'Foot'] =22*12
            RS.loc[RS['Length'].between(264,288), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '24 SYP #2'
            RS.loc[RS['Length'].between(264,288), 'Foot'] =24*12
            RS.loc[RS['Length'].between(288,312), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '26 SYP #2'
            RS.loc[RS['Length'].between(288,312), 'Foot'] =26*12
            RS.loc[RS['Length'].between(312,336), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '28 SYP #2'
            RS.loc[RS['Length'].between(312,336), 'Foot'] =28*12
            RS.loc[RS['Length'].between(336,360), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '30 SYP #2'
            RS.loc[RS['Length'].between(336,360), 'Foot'] =30*12
            RS.loc[RS['Length'].between(360,384), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '32 SYP #2'
            RS.loc[RS['Length'].between(360,384), 'Foot'] =32*12
            RS['newMAT'] = RS['newMAT'] + 'x' + RS['Length'].astype(str)
            RS1 = RS.loc[~RS['Material'].str.contains('PT')]
            RS_details = (RS1.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            RS_details = RS_details.rename(columns={'newMAT':'Material'})
            RS_details['Qty'] = (RS_details['Qty'] * RS_details['Length'])/RS_details['Foot']
            RS_details['Qty'] = RS_details['Qty'].apply(np.ceil)
            #print(RS_details," Right stud")
        
#creating separate data frame for PT SFStudRight    
        RS_details_pt = None
        RS = ll.loc[ll['Type'] == 'SFStudRight']
        if len(RS) !=0:
            RS['Length'] = RS['Length'].apply(np.ceil)
            RS['newMAT'] = RS['Material'].str.split(' ').str[0]
            RS.loc[RS['Length']<=144, 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '12 PT SYP #2'
            RS.loc[RS['Length']<=144, 'Foot'] =12*12
            RS.loc[RS['Length'].between(144,192), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '16 PT SYP #2'
            RS.loc[RS['Length'].between(144,192), 'Foot'] =16*12
            RS.loc[RS['Length'].between(192,240), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '20 PT SYP #2'
            RS.loc[RS['Length'].between(192,240), 'Foot'] =20*12
            RS.loc[RS['Length'].between(240,264), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '22 PT SYP #2'
            RS.loc[RS['Length'].between(240,264), 'Foot'] =22*12
            RS.loc[RS['Length'].between(264,288), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '24 PT SYP #2'
            RS.loc[RS['Length'].between(264,288), 'Foot'] =24*12
            RS.loc[RS['Length'].between(288,312), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '26 PT SYP #2'
            RS.loc[RS['Length'].between(288,312), 'Foot'] =26*12
            RS.loc[RS['Length'].between(312,336), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '28 PT SYP #2'
            RS.loc[RS['Length'].between(312,336), 'Foot'] =28*12
            RS.loc[RS['Length'].between(336,360), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '30 PT SYP #2'
            RS.loc[RS['Length'].between(336,360), 'Foot'] =30*12
            RS.loc[RS['Length'].between(360,384), 'From Fishbowl Inventory'] = RS['newMAT'] + 'x' + '32 PT SYP #2'
            RS.loc[RS['Length'].between(360,384), 'Foot'] =32*12
            RS['newMAT'] = RS['newMAT'] + 'x' + RS['Length'].astype(str)
            RS2 = RS.loc[RS['Material'].str.contains('PT')]
            RS_details_pt = (RS2.groupby(['newMAT','From Fishbowl Inventory','Length','Foot'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            RS_details_pt = RS_details_pt.rename(columns={'newMAT':'Material'})
            RS_details_pt['Qty'] = (RS_details_pt['Qty'] * RS_details_pt['Length'])/RS_details_pt['Foot']
            RS_details_pt['Qty'] = RS_details_pt['Qty'].apply(np.ceil)
            RS_details_pt
        
#creating separate data frame for TopPlate  
        tp_details = None          
        TP = ll.loc[ll['Type'] == 'TopPlate']
        if len(TP) != 0:
            TP['newMAT'] = TP['Material'].str.split(' ').str[0]
            TP['From Fishbowl Inventory'] = TP['newMAT'] + 'x16 SYP #2'
            TP['newMAT'] = TP['newMAT'] + 'x16'
            TP1 = TP.loc[~TP['Material'].str.contains('PT')]
            TP1['newLength'] = TP1['Length'] * TP1['Qty']
            tp_details = (TP1.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            tp_details = tp_details.rename(columns={'newMAT':'Material'})
            tp_details['Qty'] = tp_details['Qty'].apply(np.ceil)
            #print(tp_details)

#creating separate data frame for PT TopPlate 
        tp_details_pt =None       
        TP = ll.loc[ll['Type'] == 'TopPlate']
        if len(TP) != 0:
            TP['newMAT'] = TP['Material'].str.split(' ').str[0]
            TP['From Fishbowl Inventory'] = TP['newMAT'] + 'x16 PT SYP #2'
            TP['newMAT'] = TP['newMAT'] + 'x16'
            TP2 = TP.loc[TP['Material'].str.contains('PT')]
            TP2['newLength'] = TP2['Length'] * TP2['Qty']
            tp_details_pt = (TP2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            tp_details_pt = tp_details_pt.rename(columns={'newMAT':'Material'})
            tp_details_pt['Qty'] = tp_details_pt['Qty'].apply(np.ceil)
            #tp_details_pt
        
#creating separate data frame for SFTopPlate
        SFtp_details = None
        STP = ll.loc[ll['Type'] == 'SFTopPlate']
        if len(STP) != 0:
            STP['newMAT'] = STP['Material'].str.split(' ').str[0]
            STP['From Fishbowl Inventory'] = STP['newMAT'] + 'x16 SYP #2'
            STP['newMAT'] = STP['newMAT'] + 'x16'
            STP1 = STP.loc[~STP['Material'].str.contains('PT')]
            STP1['newLength'] = STP1['Length'] * STP1['Qty']
            SFtp_details =(STP1.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            SFtp_details = SFtp_details.rename(columns={'newMAT':'Material'})
            SFtp_details['Qty'] = SFtp_details['Qty'].apply(np.ceil)
            #print(SFtp_details)  

#creating separate data frame for PT SFTopPlate
        SFtp_details_pt = None
        STP = ll.loc[ll['Type'] == 'SFTopPlate']
        if len(STP) != 0:
            STP['newMAT'] = STP['Material'].str.split(' ').str[0]
            STP['From Fishbowl Inventory'] = STP['newMAT'] + 'x16 PT SYP #2'
            STP['newMAT'] = STP['newMAT'] + 'x16'
            STP2 = STP.loc[STP['Material'].str.contains('PT')]
            STP2['newLength'] = STP2['Length'] * STP2['Qty']
            SFtp_details_pt =(STP2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            SFtp_details_pt = SFtp_details_pt.rename(columns={'newMAT':'Material'})
            SFtp_details_pt['Qty'] = SFtp_details_pt['Qty'].apply(np.ceil)
        
#creating separate data frame for SFVeryTopPlate
        vtp_details = None
        VTP = ll.loc[ll['Type'] == 'SFVeryTopPlate']
        if len(VTP) != 0:
            VTP['newMAT'] = VTP['Material'].str.split(' ').str[0]
            VTP['From Fishbowl Inventory'] = VTP['newMAT'] + 'x16 SYP #2'
            VTP['newMAT'] = VTP['newMAT'] + 'x16'
            VTP1 = VTP.loc[~VTP['Material'].str.contains('PT')]
            VTP1['newLength'] = VTP1['Length'] * VTP1['Qty']
            vtp_details =(VTP1.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            vtp_details = vtp_details.rename(columns={'newMAT':'Material'})
            vtp_details['Qty'] = vtp_details['Qty'].apply(np.ceil)        
            #print(vtp_details)  

#creating separate data frame for PT SFVeryTopPlate    
        vtp_details = None    
        VTP = ll.loc[ll['Type'] == 'SFVeryTopPlate']
        if len(VTP) != 0:
            VTP['newMAT'] = VTP['Material'].str.split(' ').str[0]
            VTP['From Fishbowl Inventory'] = VTP['newMAT'] + 'x16 PT SYP #2'
            VTP['newMAT'] = VTP['newMAT'] + 'x16'
            VTP2 = VTP.loc[VTP['Material'].str.contains('PT')]
            VTP2['newLength'] = VTP2['Length'] * VTP2['Qty']
            vtp_details_pt =(VTP2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            vtp_details_pt = vtp_details_pt.rename(columns={'newMAT':'Material'})
            vtp_details_pt['Qty'] = vtp_details_pt['Qty'].apply(np.ceil)
        
#creating separate data frame for SFBlocking
        blocking_details = None
        b = ll.loc[ll['Type'] == 'SFBlocking']
        if len(b) != 0:
            b['newMAT'] = b['Material'].str.split(' ').str[0]
            b['From Fishbowl Inventory'] = b['newMAT'] + 'x16 SYP #2'
            b['newMAT'] = b['newMAT'] + 'x16'
            b1 = b.loc[~b['Material'].str.contains('PT')]
            b1['newLength'] = b1['Length'] * b1['Qty']
            blocking_details =(b1.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            blocking_details = blocking_details.rename(columns={'newMAT':'Material'})
            blocking_details['Qty'] = blocking_details['Qty'].apply(np.ceil)
            #print(blocking_details)
        
#creating separate data frame for PT SFBlocking
        blocking_details_pt = None        
        b = ll.loc[ll['Type'] == 'SFBlocking']
        if len(b) !=0:
            b['newMAT'] = b['Material'].str.split(' ').str[0]
            b['From Fishbowl Inventory'] = b['newMAT'] + 'x16 PT SYP #2'        
            b['newMAT'] = b['newMAT'] + 'x16'
            b2 = b.loc[b['Material'].str.contains('PT')]
            b2['newLength'] = b2['Length'] * b2['Qty']
            blocking_details_pt =(b2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            blocking_details_pt = blocking_details_pt.rename(columns={'newMAT':'Material'})
            blocking_details_pt['Qty'] = blocking_details_pt['Qty'].apply(np.ceil)
        
#creating separate data frame for SFBottomPlate
        bp_details_pt = None
        bp = ll.loc[ll['Type'] == 'SFBottomPlate']
        if len(bp) != 0:
            bp['newMAT'] = bp['Material'].str.split(' ').str[0]
            bp['From Fishbowl Inventory'] = bp['newMAT'] + 'x16 PT SYP #2'
            bp['newMAT'] = bp['newMAT'] + 'x16'
            bp2 = bp.loc[bp['Material'].str.contains('PT')]
            bp2['newLength'] = bp2['Length'] * bp2['Qty']
            #print(bp2)
            bp_details_pt = (bp2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            bp_details_pt = bp_details_pt.rename(columns={'newMAT':'Material'})
            bp_details_pt['Qty'] = bp_details_pt['Qty'].apply(np.ceil)
            #print("bpppppppppppp")
            #print(bp_details_pt) 
        
#creating separate data frame for PT SFBottomPlate
        bp_details = None
        bp = ll.loc[ll['Type'] == 'SFBottomPlate']
        if len(bp) != 0:
            bp['newMAT'] = bp['Material'].str.split(' ').str[0]
            bp['From Fishbowl Inventory'] = bp['newMAT'] + 'x16 SYP #2'
            bp['newMAT'] = bp['newMAT'] + 'x16'
            bp1 = bp.loc[~bp['Material'].str.contains('PT')]
            bp1['newLength'] = bp1['Length'] * bp1['Qty']
            bp_details = (bp1.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            bp_details = bp_details.rename(columns={'newMAT':'Material'})
            bp_details['Qty'] = bp_details['Qty'].apply(np.ceil)
            #print(bp_details)
            
#creating separate data frame for SFTopHeaderSill
        topheaderSill_details = None
        ths = ll.loc[ll['Type'] == 'SFTopHeaderSill']
        if len(ths) != 0:
            ths['newMAT'] = ths['Material'].str.split(' ').str[0]
            ths['From Fishbowl Inventory'] = ths['newMAT'] + 'x16 SYP #2'
            ths['newMAT'] = ths['newMAT'] + 'x16'
            ths1 = ths.loc[~ths['Material'].str.contains('PT')]
            ths1['newLength'] = ths1['Length'] * ths1['Qty']
            topheaderSill_details = (ths1.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            topheaderSill_details = topheaderSill_details.rename(columns={'newMAT':'Material'})
            topheaderSill_details['Qty'] = topheaderSill_details['Qty'].apply(np.ceil)
            #print(topheaderSill_details) 

#creating separate data frame for PT SFTopHeaderSill
        topheaderSill_details_pt = None
        ths = ll.loc[ll['Type'] == 'SFTopHeaderSill']
        if len(ths) !=0:
            ths['newMAT'] = ths['Material'].str.split(' ').str[0]
            ths['From Fishbowl Inventory'] = ths['newMAT'] + 'x16 PT SYP #2'
            ths['newMAT'] = ths['newMAT'] + 'x16'
            ths2 = ths.loc[ths['Material'].str.contains('PT')]
            ths2['newLength'] = ths2['Length'] * ths2['Qty']
            topheaderSill_details_pt = (ths2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            topheaderSill_details_pt = topheaderSill_details_pt.rename(columns={'newMAT':'Material'})
            topheaderSill_details_pt['Qty'] = topheaderSill_details_pt['Qty'].apply(np.ceil)
        
#creating separate data frame for SFBottomHeaderSill
        bottomheaderSill_details = None
        bhs = ll.loc[ll['Type'] == 'SFBottomHeaderSill']
        if len(bhs) != 0:
            bhs['newMAT'] = bhs['Material'].str.split(' ').str[0]
            bhs['From Fishbowl Inventory'] = bhs['newMAT'] + 'x16 SYP #2'        
            bhs['newMAT'] = bhs['newMAT'] + 'x16'
            bhs1 = bhs.loc[~bhs['Material'].str.contains('PT')]
            bottomheaderSill_details = (bhs1.groupby(['newMAT','From Fishbowl Inventory'])['Length'].sum()/12/16).to_frame(name='Qty').reset_index()
            bottomheaderSill_details = bottomheaderSill_details.rename(columns={'newMAT':'Material'})
            bottomheaderSill_details['Qty'] = bottomheaderSill_details['Qty'].apply(np.ceil)
            #print(bottomheaderSill_details) 
    
#creating separate data frame for PT SFBottomHeaderSill
        bottomheaderSill_details_pt = None
        bhs = ll.loc[ll['Type'] == 'SFBottomHeaderSill']
        if len(bhs) !=0:           
            bhs['newMAT'] = bhs['Material'].str.split(' ').str[0]
            bhs['From Fishbowl Inventory'] = bhs['newMAT'] + 'x16 PT SYP #2'
            bhs['newMAT'] = bhs['newMAT'] + 'x16'
            bhs2 = bhs.loc[bhs['Material'].str.contains('PT')]
            bhs2['newLength'] = bhs2['Length'] * bhs2['Qty']
            bottomheaderSill_details_pt = (bhs2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            bottomheaderSill_details_pt = bottomheaderSill_details_pt.rename(columns={'newMAT':'Material'})
            bottomheaderSill_details_pt['Qty'] = bottomheaderSill_details_pt['Qty'].apply(np.ceil)
        
#creating separate data frame for Sill 
        Sill_details = None       
        s = ll.loc[ll['Type'] == 'Sill']
        if len(s) != 0:
            s['newMAT'] = s['Material'].str.split(' ').str[0]
            s['From Fishbowl Inventory'] = s['newMAT'] + 'x16 SYP #2'
            s['newMAT'] = s['newMAT'] + 'x16'
            s1 = s.loc[~s['Material'].str.contains('PT')]
            s1['newLength'] = s1['Length'] * s1['Qty']
            Sill_details = (s1.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            Sill_details = Sill_details.rename(columns={'newMAT':'Material'})
            Sill_details['Qty'] = Sill_details['Qty'].apply(np.ceil)
            #print(Sill_details) 

#creating separate data frame for PT Sill          
        Sill_details_pt = None      
        s = ll.loc[ll['Type'] == 'Sill']
        if len(s) != 0:
            s['newMAT'] = s['Material'].str.split(' ').str[0]
            s['From Fishbowl Inventory'] = s['newMAT'] + 'x16 PT SYP #2'
            s['newMAT'] = s['newMAT'] + 'x16'
            s2 = s.loc[s['Material'].str.contains('PT')]
            s2['newLength'] = s2['Length'] * s2['Qty']
            Sill_details_pt = (s2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            Sill_details_pt = Sill_details_pt.rename(columns={'newMAT':'Material'})
            Sill_details_pt['Qty'] = Sill_details_pt['Qty'].apply(np.ceil)
        
#creating separate data frame for Header   
        missHeader = None                    
        h = ll.loc[ll['Type'] == 'Header']
        if len(h) !=0:
            h['newMAT'] = h['Material'].str.split(' ').str[0]
            h['From Fishbowl Inventory'] = h['newMAT'] + 'x16 SYP #2'
            h['newMAT'] = h['newMAT'] + 'x16'
            h2 = h.loc[(~h['Material'].str.contains('LVL'))&(~h['Material'].str.contains('PT'))]
            h2['newLength'] = h2['Length'] * h2['Qty']
            missHeader = (h2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
            missHeader=missHeader.rename(columns = {'newMAT':'Material'})            
            missHeader['Qty'] = missHeader['Qty'].apply(np.ceil)
            print(missHeader,"heDER LVL")
        
#creating separate data frame for PT Header  
        pt_header = None                     
        h = ll.loc[ll['Type'] == 'Header']
        if len(h) !=0:
            h['Length'] = h['Length'].apply(np.ceil)
            h['newMAT'] = h['Material'].str.split(' ').str[0]
            h['From Fishbowl Inventory'] = h['newMAT'] + 'x16 PT SYP #2'
            h['newMAT'] = h['newMAT'] + 'x' + h['Length'].astype(str)
            h1 = h.loc[h['Material'].str.contains('PT')]
            h1['newLength'] = h1['Length'] * h1['Qty']
            pt_header = (h1.groupby(['newMAT','From Fishbowl Inventory','newLength'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            pt_header=pt_header.rename(columns = {'newMAT':'Material'})
            pt_header['Qty'] = pt_header['newLength']/12/16
            pt_header['Qty'] = pt_header['Qty'].apply(np.ceil)
            print(pt_header,"pt header")

#creating separate data frame for LVL Header
        nonmissHeader = None                               
        h = ll.loc[ll['Type'] == 'Header']
        if len(h) !=0:
            h['Length'] = h['Length'].apply(np.ceil)
            h['newMAT'] = h['Material'].str.split(' ').str[0]
            h['From Fishbowl Inventory'] = h['newMAT'] + 'x16 SYP #2'
            h['newMAT'] = h['newMAT'] + 'x' + h['Length'].astype(str)
            h1 = h.loc[h['Material'].str.contains('LVL')]
            h1['newLength'] = h1['Length'] * h1['Qty']
            nonmissHeader = (h1.groupby(['newMAT','From Fishbowl Inventory','newLength'])['Qty'].sum()).to_frame(name='Qty').reset_index()
            nonmissHeader=nonmissHeader.rename(columns = {'newMAT':'Material'})
            nonmissHeader['Qty'] = nonmissHeader['newLength']/12/16
            nonmissHeader['Qty'] = nonmissHeader['Qty'].apply(np.ceil)
            print(nonmissHeader,"HADER MISS")
        
# =============================================================================
# #creating separate data frame for TopPlate            
#         BLOCK = ll.loc[ll['Type'] == 'SFBlocking']
#         BLOCK['newMAT'] = BLOCK['Material'].str.split(' ').str[0]
#         BLOCK['From Fishbowl Inventory'] = BLOCK['newMAT'] + 'x16 SYP #2'
#         BLOCK['newMAT'] = BLOCK['newMAT'] + 'x16'
#         BLOCK1 = BLOCK.loc[~BLOCK['Material'].str.contains('PT')]
#         BLOCK1['newLength'] = BLOCK1['Length'] * BLOCK1['Qty']
#         blocking_details = (BLOCK1.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
#         blocking_details = blocking_details.rename(columns={'newMAT':'Material'})
#         blocking_details['Qty'] = blocking_details['Qty'].apply(np.ceil)
#         #print(tp_details)
# 
# #creating separate data frame for PT TopPlate        
#         BLOCK = ll.loc[ll['Type'] == 'SFBlocking']
#         BLOCK['newMAT'] = BLOCK['Material'].str.split(' ').str[0]
#         BLOCK['From Fishbowl Inventory'] = BLOCK['newMAT'] + 'x16 PT SYP #2'
#         BLOCK['newMAT'] = BLOCK['newMAT'] + 'x16'
#         BLOCK2 = BLOCK.loc[BLOCK['Material'].str.contains('PT')]
#         BLOCK2['newLength'] = BLOCK2['Length'] * BLOCK2['Qty']
#         blocking_details_pt = (BLOCK2.groupby(['newMAT','From Fishbowl Inventory'])['newLength'].sum()/12/16).to_frame(name='Qty').reset_index()
#         blocking_details_pt = blocking_details_pt.rename(columns={'newMAT':'Material'})
#         blocking_details_pt['Qty'] = blocking_details_pt['Qty'].apply(np.ceil)
# =============================================================================
        #tp_details_pt        
        
        
#merging data frames as one list        
        stud = [Stud_details,LS_details,RS_details]
        pt_stud = [Stud_details_pt,LS_details_pt,RS_details_pt]
        king_stud = [kingStud_details]
        pt_kingstud = [kingStud_details_pt]
        cripple = [JackOverOpening_details,JackUnderOpening_details,cripple_details]#
        pt_cripple = [JackOverOpening_details_pt,JackUnderOpening_details_pt,cripple_details_pt]#
        miscellaneous = [missHeader,Sill_details,bottomheaderSill_details,topheaderSill_details,bp_details,vtp_details,SFtp_details,tp_details]
        miscellaneous_pt = [tp_details_pt,SFtp_details_pt,vtp_details_pt,bp_details_pt,topheaderSill_details_pt,bottomheaderSill_details_pt,Sill_details_pt]
        
        
        print(missHeader,"printing header")
        print("-"*20)
        print(Sill_details,"Sill")
        print("-"*20)
        print(bottomheaderSill_details,'bottom header sill')
        print("-"*20)
        print(topheaderSill_details,"top header sill")
        print("-"*20)
        print(bp_details,"bp details")
        print("-"*20)
        print(vtp_details,"vtp detail")
        print("-"*20)
        print(SFtp_details,"sf tp")
        print("-"*20)
        print(tp_details,"tp details")
        print("-"*20)
#Creating final data frames for each elements        
        stud_df = pd.concat(stud)
        Studs = stud_df.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        Studs=Studs.rename(columns = {'Material':'Material Stud'})
        Studs.index = np.arange(1,len(Studs)+1)
        #print(Studs)
        
        pt_stud_df = pd.concat(pt_stud)
        Pt_studs = pt_stud_df.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        Pt_studs=Pt_studs.rename(columns = {'Material':'Material PT Stud'})
        Pt_studs.index = np.arange(1,len(Pt_studs)+1)
        
        
        #print(Pt_studs)
        
        cripple_df = pd.concat(cripple)
        Cripple = cripple_df.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        Cripple=Cripple.rename(columns = {'Material':'Material Cripple'})
        Cripple.index = np.arange(1,len(Cripple)+1)
        
        #print(Cripple)
        
        pt_cripple_df = pd.concat(pt_cripple)
        Pt_cripple = pt_cripple_df.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        Pt_cripple=Pt_cripple.rename(columns = {'Material':'Material PT Cripple'})
        Pt_cripple.index = np.arange(1,len(Pt_cripple)+1)
        
        king = pd.concat(king_stud)
        King = king.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        King=King.rename(columns = {'Material':'Material KingStud'})
        King.index = np.arange(1,len(King)+1)
        
        kingstudPt =pd.concat(pt_kingstud) 
        King_pt = kingstudPt.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        King_pt=King_pt.rename(columns = {'Material':'Material PT KingStud'})
        King_pt.index = np.arange(1,len(King_pt)+1)
        
        Jack = Beam_details.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        Jack=Jack.rename(columns = {'Material':'Material Jack'})
        Jack.index = np.arange(1,len(Jack)+1)
        
        Jack_pt = Beam_details_pt.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        Jack_pt=Jack_pt.rename(columns = {'Material':'Material PT Jack'})
        Jack_pt.index = np.arange(1,len(Jack_pt)+1)
        
        headlvl = nonmissHeader.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        headlvl=headlvl.rename(columns = {'Material':'Material LVL Header'})
        headlvl.index = np.arange(1,len(headlvl)+1)
        
        header_pt = pt_header.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        header_pt=header_pt.rename(columns = {'Material':'Material PT Header'})
        header_pt.index = np.arange(1,len(header_pt)+1)
        
        
        blockings = blocking_details.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        blockings=blockings.rename(columns = {'Material':'Material SFBlockings'})
        blockings.index = np.arange(1,len(blockings)+1)
        
        blockings_pt = blocking_details_pt.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        blockings_pt=blockings_pt.rename(columns = {'Material':'Material PT SFBlockings'})
        blockings_pt.index = np.arange(1,len(blockings_pt)+1)
        
        post = Post_details.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        post=post.rename(columns = {'Material':'Material Post'})
        post.index = np.arange(1,len(post)+1)
        
        post_pt = Post_details_pt.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        post_pt=post_pt.rename(columns = {'Material':'Material PT Post'})
        post_pt.index = np.arange(1,len(post_pt)+1)
        
        post_lvl = Post_details_lvl.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        post_lvl=post_lvl.rename(columns = {'Material':'Material LVL Post'})
        post_lvl.index = np.arange(1,len(post_lvl)+1)
        
        #print(post,post_pt,post_lvl)
        miss =pd.concat(miscellaneous) 
        Miss = miss.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        Miss=Miss.rename(columns = {'Material':'Material Miscellaneous'})
        Miss.index = np.arange(1,len(Miss)+1)
        
        print(Miss)
        
        miss_pt =pd.concat(miscellaneous_pt) 
        Miss_pt = miss_pt.groupby(['Material','From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        Miss_pt=Miss_pt.rename(columns = {'Material':'Material PT Miscellaneous'})
        Miss_pt.index = np.arange(1,len(Miss_pt)+1)
        #print(Miss_pt)
        
#Functions to add multiple data frames into same excelsheet        
        def multiple_dfs(df_list, sheets, file_name, spaces):
            writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
            row = 1
            for dataframe in df_list:
                dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
                row = row + len(dataframe.index) + spaces + 1
            writer.save()
        
#removing dataframes which has 0 elements
        dfs = [x for x in [Studs,Pt_studs,Cripple,Pt_cripple,King,King_pt,Jack,Jack_pt,headlvl,header_pt,post,post_pt,post_lvl,Miss,Miss_pt,blockings,blockings_pt] if len(x)!=0]#,
        
        
        #print(dfs)
        # run function
        multiple_dfs(dfs, 'Validation', output+'\\Cutlist.xlsx', 1)
        
        newDfs = pd.concat(dfs)
        newDfs = newDfs.reset_index()
        
        twoColDf = newDfs[['From Fishbowl Inventory','Qty']]
        twoColDf = twoColDf.replace(to_replace ="X",value="x",regex=True)
        twoColDf = twoColDf.sort_values(by=['From Fishbowl Inventory'])
        twoColDf = twoColDf.reset_index()
        twoColDf= twoColDf.drop('index',axis=1)
        
        total = twoColDf.groupby(['From Fishbowl Inventory'])['Qty'].sum().to_frame(name='Qty').reset_index()
        fishbowl = total['From Fishbowl Inventory'].to_list()
        new_fishbowl = [i+'Total' for i in fishbowl]
        
        Qty = total['Qty'].to_list()
        
        df_list = [split for _, split in twoColDf.groupby(['From Fishbowl Inventory'])]
        count=0
        long_df = []
        for x in df_list:
            x.loc[len(x.index)] = [new_fishbowl[count],Qty[count]]
            count+=1
            long_df.append(x)
        
        
        sheet1 = pd.concat(long_df)
        sheet1 = sheet1.reset_index()
        sheet1= sheet1.drop('index',axis=1)
        sheet1
        
        
        
        writer = pd.ExcelWriter(output+'\\Cutlist.xlsx', engine='openpyxl', mode='a')
        writer.book = load_workbook(output+'\\Cutlist.xlsx')
        sheet1.to_excel(writer,sheet_name='d1')
        writer.save()
        writer.close()
        
        writer = pd.ExcelWriter(output+'\\Cutlist.xlsx', engine='openpyxl', mode='a')
        writer.book = load_workbook(output+'\\Cutlist.xlsx')
        total.to_excel(writer,sheet_name='d2')
        writer.save()
        writer.close()
        
        completed = Label(canvas , text='Completed !!',font=('Arial',20),foreground="#FFDF00",background='#010203')
        completed.place(x=100,y=250)
        completed.after(6000, lambda:  completed.destroy() )
# =============================================================================
#         canvas.create_text(165, 350, text='COMPLETED',
#                            font=('Cambria', 16, 'bold'), fill='#FFDF00')
# =============================================================================
        
        
btn3 = Button(canvas, text='CREATE', font=('Cambria', 10, 'bold'),
              command=generate, bg='#00003f', fg='#FFDF00')

canvas.create_window(90, 180, anchor='e', window=btn3, height=30, width=70)
troot.mainloop()